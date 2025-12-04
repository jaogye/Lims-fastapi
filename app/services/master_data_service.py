"""
Master data service module.

This module provides services for importing and exporting master data via Excel files.
Handles data validation, transformation, and bulk operations for products, qualities,
specifications, and other reference data. Migrated from MATLAB master data management system.
"""

from sqlalchemy.orm import Session
from fastapi import UploadFile
import pandas as pd
import tempfile
import os
from typing import List, Dict, Any, Optional
from tempfile import TemporaryDirectory
import logging
from openpyxl.utils import get_column_letter
from sqlalchemy import  text
from ..services.view import saveView, getView

logger = logging.getLogger(__name__)


table_types = ["products", "qualities", "variables", "holidays", "sample_points",
                       "spec-client", "spec-gen", "samplematrix", "maps" ]

class MasterDataQuery:
    """
    Service class for querying and exporting master data.

    Provides methods for retrieving master data and exporting to Excel
    with auto-sized columns for better readability.

    Attributes:
        db (Session): SQLAlchemy database session.
    """    
    def __init__(self, db: Session):
        self.db = db
              
    def _auto_size_columns(self, worksheet, df: pd.DataFrame):
        """Auto-size Excel columns based on content"""
        for idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            # Calculate max length for column
            # Check column header length
            max_length = len(str(col))
            
            # Check data lengths
            if col in df.columns:
                try:
                    max_data_length = df[col].astype(str).str.len().max()
                    max_length = max(max_length, max_data_length)
                except:
                    pass
            
            # Add some padding and set a reasonable max width
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    async def export_to_excel(self, table_type: str) -> str:
        """Export master data to Excel file with auto-sized columns"""
        logger.info(f"Exporting {table_type} to Excel")
        with TemporaryDirectory() as temp_dir:
            filename = f"{table_type}.xlsx"
            excel_path = os.path.join(temp_dir, filename)
            try:
                if table_type in table_types:
                    logger.info(f"Fetching all {table_type}")
                    data = getView(self.db, table_type)
                else:
                   logger.error(f"Unsupported table type: {table_type}. Valid table types: {table_types}")
                   raise ValueError(f"Unsupported table type: {table_type}. Valid table types: {table_types}")

                df = pd.DataFrame(data)
                # Export to Excel with auto-sizing
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=table_type, index=False)

                    # Auto-size columns
                    worksheet = writer.sheets[table_type]
                    self._auto_size_columns(worksheet, df)

                logger.info(f"Successfully exported {len(df)} rows to {excel_path}")

                # Copy file to persistent location before temp dir is deleted
                persistent_dir = tempfile.gettempdir()
                persistent_path = os.path.join(persistent_dir, filename)

                import shutil
                shutil.copy(excel_path, persistent_path)

                return persistent_path
            except Exception as e:
                logger.error(f"Failed to export {table_type} to Excel: {e}")
                raise

    async def get_products(self):
        """Get list of all products"""
        query = text("SELECT id, name, bruto FROM product ORDER BY name")
        result = self.db.execute(query)
        products = []
        for row in result:
            products.append({
                "id": row[0],
                "name": row[1],
                "bruto": row[2]
            })
        return products

    async def get_qualities(self):
        """Get list of all qualities"""
        query = text("SELECT id, name, long_name FROM quality ORDER BY name")
        result = self.db.execute(query)
        qualities = []
        for row in result:
            qualities.append({
                "id": row[0],
                "name": row[1],
                "long_name": row[2]
            })
        return qualities

    async def get_sample_points(self):
        """Get list of all sample points"""
        query = text("SELECT id, name FROM samplepoint ORDER BY name")
        result = self.db.execute(query)
        sample_points = []
        for row in result:
            sample_points.append({
                "id": row[0],
                "name": row[1]
            })
        return sample_points

    async def get_variables(self):
        """Get list of all variables"""
        query = text("SELECT id, short_name, test, element, unit, ord FROM variable ORDER BY ord, short_name")
        result = self.db.execute(query)
        variables = []
        for row in result:
            variables.append({
                "id": row[0],
                "short_name": row[1],
                "test": row[2],
                "element": row[3],
                "unit": row[4],
                "ord": row[5]
            })
        return variables

    async def get_qualities_by_product(self, product_id: int):
        """Get list of qualities filtered by product using spec table"""
        query = text("""
            SELECT DISTINCT q.id, q.name, q.long_name
            FROM quality q, spec s
            WHERE s.type_spec='GEN'
              AND s.quality_id = q.id
              AND s.product_id = :product_id
            ORDER BY q.name
        """)
        result = self.db.execute(query, {"product_id": product_id})
        qualities = []
        for row in result:
            qualities.append({
                "id": row[0],
                "name": row[1],
                "long_name": row[2]
            })
        return qualities

    async def get_spec_id(self, product_id: int, quality_id: int):
        """Get spec_id from product_id and quality_id"""
        query = text("""
            SELECT id
            FROM spec
            WHERE type_spec='GEN'
              AND product_id = :product_id
              AND quality_id = :quality_id
        """)
        result = self.db.execute(query, {"product_id": product_id, "quality_id": quality_id})
        row = result.fetchone()
        if row:
            return row[0]
        return None

    async def get_sample_points_by_product_quality(self, product_id: int, quality_id: int):
        """Get sample points filtered by product and quality using samplematrix"""
        query = text("""
            SELECT DISTINCT sp.id, sp.name
            FROM samplepoint sp, samplematrix sm
            WHERE sm.samplepoint_id = sp.id
              AND sm.product_id = :product_id
              AND sm.quality_id = :quality_id
            ORDER BY sp.name
        """)
        result = self.db.execute(query, {"product_id": product_id, "quality_id": quality_id})
        sample_points = []
        for row in result:
            sample_points.append({
                "id": row[0],
                "name": row[1]
            })
        return sample_points



# Class that stores the DML methods
class MasterDataService:
    """
    Service class for importing and modifying master data.

    Handles Excel file uploads, data validation, and bulk insert/update/delete
    operations for master data tables. Provides error handling and reporting
    for failed records.

    Attributes:
        db (Session): SQLAlchemy database session.
    """

    def __init__(self, db: Session):
        """
        Initialize the master data service.

        Args:
            db (Session): SQLAlchemy database session.
        """
        self.db = db
                           
    async def import_from_excel(
        self,
        file: UploadFile,
        table_type: str
    ) -> Dict[str, Any]:
        """Import master data from Excel file"""
        logger.info(f"Importing {table_type} from Excel file: {file.filename}")
        with TemporaryDirectory() as temp_dir:
            temp_path = os.path.join(temp_dir, file.filename)
            try:
                with open(temp_path, "wb") as temp_file:
                    content = await file.read()
                    temp_file.write(content)
                df = pd.read_excel(temp_path, sheet_name=0)
                if table_type in table_types:
                    msgerror, stat, pendingdata= saveView(table_type, self.db, df)

                    error_file_url = None
                    # If there are errors, save the non-processed rows to an Excel file
                    if len(msgerror) > 0 and not pendingdata.empty:
                        error_file_url = self._save_error_rows(pendingdata, table_type)
                        logger.info(f"Saved {len(pendingdata)} error rows to {error_file_url}")

                    # Return errors and error file info without raising exception
                    return {
                        "processed": stat,
                        "errors": msgerror,
                        "pendingdata": pendingdata.to_dict('records'), 
                        "error_file": error_file_url,
                        "has_errors": len(msgerror) > 0
                    }
                else:
                   logger.error(f"Unsupported table type: {table_type}. Valid table types: {table_types}")
                   raise ValueError(f"Unsupported table type: {table_type}. Valid table types: {table_types}")
            except Exception as e:
                logger.error(f"Failed to import from Excel: {e}")
                raise

        return {"processed":stat, "errors": [], "pendingdata": pendingdata, "error_file": None, "has_errors": False}

    def _save_error_rows(self, pendingdata: pd.DataFrame, table_type: str) -> str:
        """Save non-processed rows to an Excel file and return the file path"""
        import uuid
        from datetime import datetime

        # Generate unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_id = str(uuid.uuid4())[:8]
        filename = f"{table_type}_errors_{timestamp}_{unique_id}.xlsx"

        # Save to persistent temp directory
        persistent_dir = tempfile.gettempdir()
        error_file_path = os.path.join(persistent_dir, filename)

        # Export error rows to Excel with auto-sizing
        with pd.ExcelWriter(error_file_path, engine='openpyxl') as writer:
            pendingdata.to_excel(writer, sheet_name='Errors', index=False)

            # Auto-size columns (reuse the method from MasterDataQuery)
            worksheet = writer.sheets['Errors']
            for idx, col in enumerate(pendingdata.columns, 1):
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(idx)
                max_length = len(str(col))

                if col in pendingdata.columns:
                    try:
                        max_data_length = pendingdata[col].astype(str).str.len().max()
                        max_length = max(max_length, max_data_length)
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Saved error file to: {error_file_path}")
        return filename                

